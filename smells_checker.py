#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import sys

import openpyxl

from collections import Counter


"""
    Given a xlsx worksheet, read the work_books and gather the informations about the smells
"""
def count_smels(work_sheet):
    
    #store the results
    smells_results = {}

    #target workbooks -> should extract information only os this workbooks
    target_workbooks = ['_ArchSMells', '_AbsSMells', '_EncSMells', '_ModSMells', '_HieSMells', '_ImpSmells', '_CodeClones']
    
    #headers names
    headers_list = ['Architecture smell', 'Design smell', 'Implementation smell', 'Clone-set Serial No']

    #a list with the name of the work_books
    work_books_list = work_sheet.sheetnames
    
    for work_book in work_books_list:
        
        #avoid unnecessary work books
        if not any(x in work_book for x in target_workbooks): #check sheets the contains any of the listed words
            continue

        #get a valid work book
        sheet = work_sheet[work_book]

        for col in sheet.iter_cols(min_row=1, max_col=1):
            for cell in col:

                #get the value of the cell 
                smell_name = cell.value

                #avoid headers
                if smell_name in headers_list:
                    continue               
                
                #handle code duplucation identifier
                if isinstance(smell_name, int) or smell_name is None:
                    smell_name = 'Clone'
                
                #first occurrence
                if not smell_name in smells_results.keys():

                    smells_results[smell_name] = 1
                
                #increment the number of occurrence
                else:
                    counter = smells_results[smell_name]
                    counter += 1
                    smells_results[smell_name] = counter                    

    return smells_results
                
if __name__ == '__main__':

    #gather results
    results = dict()

    #base dir
    base_dir = os.getcwd()
    
    #Check python version
    if sys.version_info[0] < 3:
        raise Exception("Python 3 or a more recent version is required.")
        #run -> python3 smells_checker.py >> text.log    

    repositories_path  = base_dir + "/repositories/smells/vr/"

    os.chdir(repositories_path)

    for root, dirs, files in os.walk(".", topdown=False):
        for filename in files:

            if '.xlsx' in filename:               
                
                result = dict()
                           
                #open a xlsx file
                work_sheet = openpyxl.load_workbook(filename)
                result = count_smels(work_sheet)

                #print("{}   :   {}".format(filename, result)) #check individual iterations

                #merge the results
                results = dict(Counter(results) + Counter(result))
    
    #count the smeels
    print(results)