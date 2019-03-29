#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import sys

import openpyxl
import pandas as pd

#extra
import xlrd
import datetime
from itertools import chain


#needs to install Gnumeric spreadsheet program to convert the files
def xlsx_to_csv(target_dir):

	for file in os.listdir(target_dir):

		#read only xlsx files, ignore other files
		if ".xlsx" not in file and 'xls' not in file:
			continue

		if 'xlsx' in file:
			new_file = file[:-5] + '.csv'
		else:
			new_file = file[:-4] + '.csv' 
		
		#print statment
		statment = 'ssconvert {} {}'.format(file,new_file)
		os.system(statment)



"""
	Convert xls files to xlsx files. Credits -> https://stackoverflow.com/questions/9918646/how-to-convert-xls-to-xlsx
"""
def xls_to_xlsx(*args, **kw):
	"""
	open and convert an XLS file to openpyxl.workbook.Workbook
	----------
	@param args: args for xlrd.open_workbook
	@param kw: kwargs for xlrd.open_workbook
	@return: openpyxl.workbook.Workbook
	"""
	
	book_xls = xlrd.open_workbook(*args, formatting_info=True, ragged_rows=True, **kw)
	book_xlsx = openpyxl.workbook.Workbook()

	sheet_names = book_xls.sheet_names()
	for sheet_index in range(len(sheet_names)):
		sheet_xls = book_xls.sheet_by_name(sheet_names[sheet_index])
		if sheet_index == 0:
			sheet_xlsx = book_xlsx.active
			sheet_xlsx.title = sheet_names[sheet_index]
		else:
			sheet_xlsx = book_xlsx.create_sheet(title=sheet_names[sheet_index])
		for crange in sheet_xls.merged_cells:
			rlo, rhi, clo, chi = crange
			sheet_xlsx.merge_cells(start_row=rlo + 1, end_row=rhi,
			start_column=clo + 1, end_column=chi,)

		def _get_xlrd_cell_value(cell):
			value = cell.value
			if cell.ctype == xlrd.XL_CELL_DATE:
				datetime_tup = xlrd.xldate_as_tuple(value,0)	
				if datetime_tup[0:3] == (0, 0, 0):   # time format without date
					value = datetime.time(*datetime_tup[3:])
				else:
					value = datetime.datetime(*datetime_tup)
			return value

		for row in range(sheet_xls.nrows):
			sheet_xlsx.append((
				_get_xlrd_cell_value(cell)
				for cell in sheet_xls.row_slice(row, end_colx=sheet_xls.row_len(row))
			))
	
	filename = str(*args)
	filename = filename.replace('xls', 'xlsx')
	book_xlsx.save(filename)


"""
	Some projects have duplicated files, so designite create double entries (duplicated lines) when calculate/export smells to csv.
	Create this rotine to remove the duplicated lines in every csv file.
	https://stackoverflow.com/questions/15741564/removing-duplicate-rows-from-a-csv-file-using-a-python-script  
"""
def remove_duplicated_lines(afile):

	#create a result dir to storage the new files
	result_dir = os.getcwd() + '/new_csv/'
	
	#simple dir check
	try:
		#verifies if its a valid dir	
		os.stat(result_dir)
	except:
		#since the dir doesn't exist, create it!
		os.mkdir(result_dir)
	
	#result file will have the same name
	new_file = result_dir + afile

	#routine
	with open(afile,'r') as in_file, open(new_file,'w') as out_file:
		seen = set() # set for fast O(1) amortized lookup
		for line in in_file:
			if line in seen: continue # skip duplicate

			seen.add(line)
			out_file.write(line)



"""
	Given a dir that contains the results of designite about code smells into csv format, read all csv files and retrive the name of all
	classes analyzed  
"""
def get_classe_names_from_csv_dir(csv_dir):
	
	#enter into the dir
	try:
		os.chdir(csv_dir)
	except Exception as e:
		print e
		print "E: Invalid directory. Ending execution. (get_classe_names_from_csv_dir)"
		sys.exit(-1)

	#will store the results into a dict
	results = {}

	# traverse root directory, and list directories as dirs and files as files
	for root, dirs, files in os.walk("."):
		
		#check every file		
		for file_name in files:
			
			#check if it's a csv file
			if file_name[-4:] == ".csv":
				
				#add the complet path to the file
				target_file = csv_dir + '/' + file_name

				#create a dataframe with the information of the csv
				data_frame = pd.read_csv(target_file)

				#get the name of the classes
				classes = list(data_frame['Type'])
				
				#creat a dict wich the key is the name of the csv and the value is a list with the classes present into that csv file
				results[file_name[:-4]] = classes			
	
	#just return the results
	return results




"""
	Given a directory check all files and search for C# test classes 
"""
def search_all_test_class(projects_dir):


	#will store the results.:: Format -> Name of the file -> path of the file
	results = {}

	#enter into the dir
	try:
		os.chdir(projects_dir)
	except Exception as e:
		print "E: Invalid directory. Ending execution. (search_all_test_class)"
		sys.exit(-1)

	# traverse root directory, and list directories as dirs and files as files
	for root, dirs, files in os.walk("."):

		#print the base path
		path = root.split(os.sep)
		#print (len(path) - 1) * '+++', os.path.basename(root)
		
		#check every file		
		for file_name in files:

			#check if it's a C# file
			if file_name[-3:] == ".cs":
			
				#add the complet path to the file
				target_file = os.getcwd() + '/' + root + '/' + file_name
				
				#open the .cs file
				with open(target_file, 'r') as afile:

					#convert the content of the file into a string
					afile_content = afile.read().replace('\n', '')

					#test evidences will identify whether afile has test content or not
					test_evidence = "[TestFixture]"
					
					#check the content of afile
					if test_evidence in afile_content:

						#save a dict with all classes that contains C# tests
						results[file_name] = target_file

	#just return the results
	return results
							

"""
	Will evaluate all classes of the experiment individually in order to verify wheter that class was properly tested or not
	We wanna know all classes that were properly tested
"""
def find_all_classes_tested(projects_dir, csv_dir):

	#will store the results
	results = {}

	#first get all classes used in the experiment

	#get the name of all classes used in the experiment:. Format:: key = project name; values = list of the classes.: eg: project -> [v, a, l, u, e, s]
	experimented_classes_dict = get_classe_names_from_csv_dir(csv_dir)
	
	#get only the name of the classes used
	experiment_classes = list(chain(*experimented_classes_dict.values()))

	#remove duplicates
	experiment_classes = list(set(experiment_classes))
	
	#remove possible nans from the list
	experiment_classes = [x for x in experiment_classes if str(x) != 'nan']
	
	#now get all classes of test of the repositories
	unit_test_classes  = search_all_test_class(projects_dir)

	#key is the name of the classe and value is the path of the class
	for one_unit_test_class, path_of_unit_test_class in unit_test_classes.items():

		#open every unit test class
		with open(path_of_unit_test_class, 'r') as target_test_class:

			#convert the content of the unit test class into a string
			target_test_class_content = target_test_class.read().replace('\n', '')

			#verify every class that was used in the experiment
			for one_experimented_class in experiment_classes:

				#this occurs when the class is checked against itself
				if one_experimented_class == one_unit_test_class:
					continue

				#check whether that class was tested or not 
				if one_experimented_class in target_test_class_content:

					#first occurrence 
					if one_unit_test_class not in results.keys():
						results[one_unit_test_class] = [one_experimented_class]

					#avoid duplication
					if one_experimented_class not in results[one_unit_test_class]:

						#add new occurrence
						results[one_unit_test_class] = results[one_unit_test_class] + [one_experimented_class] 

	#remove duplicated
	tested_classes = list(set(chain(*results.values())))
	test_classes =   unit_test_classes.values()
	
	print "Number of classes of test: ", len(test_classes)
	print "Number of classes tested: ",  len(tested_classes)

	
	#will store the results into a pandas' data_frame
	data_frame = pd.DataFrame(tested_classes, columns=["tested_classes"])

	return data_frame
	


#used to format some data before run the other scripts
if __name__ == "__main__":

	#this solution relies on python 2.X

	#configurations
	target_dir = '.'
	os.chdir(target_dir)

	#possible methods
	print "1 - Remove duplicates from CSV"
	print "2 - Convert XLSX -> CSV"
	print "3 - Convert XLS  -> XLSX"
	print "4 - Find C# tested classes"

	#read the option
	switch = int(raw_input("Chose an method: "))
	
	#execute according to the option
	if switch not in [1, 2, 3, 4]:
		print "W: Invalid option!"

	if switch == 1:

		#run for each file from the dir
		for root, dirs, files in os.walk(".", topdown=False):
			for file_name in files:

				#avoid other files				
				if ".csv" not in file_name:
					continue

				#execute on csv
				remove_duplicated_lines(file_name)


	if switch == 2:

		#simple do the conversion for all files in the dir (gnumeric tool required)
		xlsx_to_csv(target_dir)

	if switch == 3:

		#read of files from dir
		for root, dirs, files in os.walk(".", topdown=False):
			for file_name in files:				
				
				#avoid non xls files
				if ".xls" not in file_name:
					continue
				
				#do the conversion 
				xls_to_xlsx(file_name)

	if switch == 4:

		base_dir = os.getcwd()

		csv_dir      = base_dir + "/results/full_results/metrics/non-vr" #dir to the results of full smells
		projects_dir = base_dir + "/repositories/source"        #check the repositories info to find the links to the repos source code

		print "I: This operation can take sometime. In the end will save the results into a csv file."

		#can take sometime.. 
		data_frame = find_all_classes_tested(projects_dir, csv_dir)

		#change to the correct place
		os.chdir(base_dir + "/results/sampled_results/")

		#save the data frame into a csv file to futher use
		data_frame.to_csv('tested_classes.csv', index=False)